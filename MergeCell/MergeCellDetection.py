# -*- coding: UTF-8 -*-
"""
@Intorduce：可以检测合并单元格
@Project ：_init_.py 
@File ：MergeCellDetection.py
@Author ：小小小松
@Date ：2023/4/30 22:51
"""
import os
import openpyxl
import pandas as pd
from data import DataPath


def is_merged_cell(cell):
    """
    判断单元格是否为合并单元格
    """
    if cell.coordinate in cell.parent.merged_cells:
        return True
    return False


def HasMergedCell(path):
    IsMerged=False
    workbook = openpyxl.load_workbook(path)
    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        df = pd.DataFrame(sheet.values)
        merged_cells = sheet.merged_cells

        # 遍历所有单元格
        for index, row in df.iterrows():
            for column in df.columns:
                cell_value = row[column]
                cell = sheet.cell(row=index + 1, column=column + 1)
                if is_merged_cell(cell) and cell.coordinate in merged_cells:
                    # print(f"合并单元格：({index + 1}, {column + 1}),值为：{cell_value}")
                    IsMerged = True
                # if pd.isna(cell_value):
                #     print(f"空单元格：({index + 1}, {column + 1})")
                # elif is_merged_cell(cell) and cell.coordinate in merged_cells:
                #     print(f"合并单元格：({index + 1}, {column + 1}),值为：{cell_value}")
                #     IsMerged=True
                # else:
                #     print(f"非空单元格：({index + 1}, {column + 1})，值为：{cell_value}")
    # 输出处理后的DataFrame
    return IsMerged


if __name__ == '__main__':
    excels = ['专家经验树.xlsx', '文本挖掘树_.xlsx', "机器学习特征树_.xlsx", '融合树_.xlsx']

    path = os.path.join(DataPath, excels[0])
    HasMergedCell(path)

