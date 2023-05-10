# -*- coding: UTF-8 -*-
"""
@Intorduce：拆分一列中存在的合并单元格
@Project ：_init_.py 
@File ：MergeCellSplit.py
@Author ：小小小松
@Date ：2023/4/30 23:27
"""
import os

import pandas as pd

from data import DataPath
import openpyxl


def MergeCellSplit(path):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(path)
    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        pd.DataFrame(worksheet.values)

        m_list = worksheet.merged_cells
        # 合并单元格的位置信息，可迭代对象（单个是一个'openpyxl.worksheet.cell_range.CellRange'对象），print后就是excel坐标信息
        cr = []
        for m_area in m_list:
            # 合并单元格的起始行坐标、终止行坐标。。。。，
            r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
            # 纵向合并单元格的位置信息提取出
            if r2 - r1 > 0:
                cr.append((r1, r2, c1, c2))
                # print('符合条件%s' % str(m_area))
        # 这里注意需要把合并单元格的信息提取出再拆分
        for r in cr:
            worksheet.unmerge_cells(start_row=r[0], end_row=r[1],
                                    start_column=r[2], end_column=r[3])
            for row in range(r[0], r[1]):
                worksheet.cell(row=row + 1, column=r[3], value=worksheet.cell(r[0], r[3]).value)
    workbook.save(path)


if __name__ == '__main__':
    path=os.path.join(DataPath, "Data.xlsx")
    # excels = ['专家经验树.xlsx', '文本挖掘树_.xlsx', "机器学习特征树_.xlsx", '融合树_.xlsx']
    # path = os.path.join(DataPath, excels[0])
    MergeCellSplit(path)